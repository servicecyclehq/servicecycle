#!/usr/bin/env python3
"""
audit_xlsx.py - consistency QA for generated Excel workbooks.

Verifies every data-bearing sheet carries the canonical ServiceCycle styling so
no export ships as a raw, unstyled grid. Per sheet it checks:
  - BRANDED: a branded fill appears in the top rows (petrol header, ink masthead,
             or KPI/zebra wash) - i.e. the sheet is styled, not a raw dump.
  - FREEZE:  freeze panes are set (header stays visible while scrolling).
  - TABCOLOR: the sheet tab is colored (WARN only).

Non-zero exit if any ERROR, so it can gate a deploy. Optionally renders a
contact sheet per workbook via LibreOffice (--png-dir DIR).

Requires: openpyxl. Usage:
  python3 audit_xlsx.py <file.xlsx> [more.xlsx ...] [--png-dir DIR]
"""
import sys, os, subprocess
import openpyxl

# Canonical brand fills (ARGB, upper). Mirrors lib/xlsxStyle BRAND.
BRAND_FILLS = {
    'FF0D4F6E',          # petrol header
    'FF0A0D12', 'FF111827',  # ink masthead
    'FFE9F1F5',          # accentLt KPI wash
    'FFF7F9FB', 'FFF4F7FB',  # zebra band
    'FFAFCFDD',          # data-bar / light petrol
    'FFE7F5EC', 'FFFCF1E2', 'FFFBEAEA',  # status chips
}
TOP_ROWS = 8   # branded styling must appear within the first N rows


def fill_rgb(cell):
    try:
        f = cell.fill
        if f is not None and f.patternType == 'solid':
            rgb = f.fgColor.rgb
            if isinstance(rgb, str):
                return rgb.upper()
    except Exception:
        pass
    return None


def audit_sheet(ws):
    findings = []
    maxr, maxc = ws.max_row or 0, ws.max_column or 0
    if maxr < 2 or maxc < 1:
        return findings  # effectively empty; nothing to brand

    branded = False
    for r in range(1, min(TOP_ROWS, maxr) + 1):
        for c in range(1, min(maxc, 40) + 1):
            if fill_rgb(ws.cell(row=r, column=c)) in BRAND_FILLS:
                branded = True
                break
        if branded:
            break
    if not branded:
        findings.append(('ERROR', 'no branded styling in the top rows (raw/unstyled export?)'))

    if not ws.freeze_panes:
        findings.append(('ERROR', 'no freeze panes (header will scroll away)'))

    if not ws.sheet_properties.tabColor:
        findings.append(('WARN', 'no tab color'))

    return findings


def contact_sheet(xlsx, png_dir):
    try:
        subprocess.run(['soffice', '--headless', '--convert-to', 'pdf', '--outdir', png_dir, xlsx],
                       check=True, timeout=60, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
        pdf = os.path.join(png_dir, os.path.basename(xlsx).replace('.xlsx', '.pdf'))
        import fitz
        doc = fitz.open(pdf)
        for i, p in enumerate(doc):
            p.get_pixmap(dpi=130).save(os.path.join(png_dir, os.path.basename(xlsx).replace('.xlsx', f'_p{i+1}.png')))
        return doc.page_count
    except Exception as e:
        return 'render failed: %s' % e


def main():
    args, png_dir, argv, i = [], None, sys.argv[1:], 0
    while i < len(argv):
        a = argv[i]
        if a == '--png-dir':
            png_dir = argv[i + 1] if i + 1 < len(argv) else None
            i += 2
            continue
        if a.startswith('--'):
            i += 1
            continue
        args.append(a)
        i += 1

    total_err = total_warn = 0
    for f in args:
        wb = openpyxl.load_workbook(f)
        print('\n=== %s  (%d sheets) ===' % (os.path.basename(f), len(wb.sheetnames)))
        ferr = 0
        for name in wb.sheetnames:
            for level, msg in audit_sheet(wb[name]):
                if level == 'ERROR':
                    ferr += 1
                    print('  ERROR  [%s]: %s' % (name, msg))
                else:
                    total_warn += 1
                    print('  warn   [%s]: %s' % (name, msg))
        if ferr == 0:
            print('  OK - every data sheet branded + frozen')
        total_err += ferr
        if png_dir:
            print('  contact sheet pages: %s' % contact_sheet(f, png_dir))
    print('\nTOTAL: %d ERROR, %d warn' % (total_err, total_warn))
    sys.exit(1 if total_err else 0)


if __name__ == '__main__':
    main()
